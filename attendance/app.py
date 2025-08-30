# app.py
# Attendance taker web app with recheck absentees feature.
# Requirements: flask, pandas, openpyxl
# pip install flask pandas openpyxl

import pandas as pd
from datetime import datetime
import threading
import os
import logging
import time
import webbrowser

from flask import Flask, render_template_string, request, jsonify
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')

app = Flask(__name__)

# --- Read Excel filename from text file ---
FILENAME_TXT = "input_filename.txt"
if os.path.exists(FILENAME_TXT):
    with open(FILENAME_TXT, "r", encoding="utf-8") as f:
        INPUT_EXCEL_FILE = f.read().strip()
else:
    INPUT_EXCEL_FILE = None

save_lock = threading.Lock()


def read_and_index_df():
    """Read the Excel file and return DataFrame indexed by OrgDefinedId."""
    if not INPUT_EXCEL_FILE or not os.path.exists(INPUT_EXCEL_FILE):
        return None, f"Input file '{INPUT_EXCEL_FILE}' not found."

    try:
        df = pd.read_excel(INPUT_EXCEL_FILE, engine='openpyxl', dtype=str)
    except Exception as e:
        return None, f"Error reading Excel file: {e}"

    if 'OrgDefinedId' not in df.columns:
        return None, "Required column 'OrgDefinedId' not found in the Excel file."

    df['OrgDefinedId'] = df['OrgDefinedId'].astype(str).str.strip()
    df.set_index('OrgDefinedId', inplace=True, drop=False)
    return df, None


@app.route('/')
def home():
    return HTML_TEMPLATE


@app.route('/get_students')
def get_students():
    df, err = read_and_index_df()
    if err:
        return jsonify({'error': err})

    students_list = []
    for idx, row in df.iterrows():
        first = row.get('First Name') or row.get('First') or row.get('GivenName') or ''
        last = row.get('Last Name') or row.get('Last') or row.get('Surname') or ''
        name = f"{first} {last}".strip() or f"Student {idx}"
        students_list.append({'id': str(idx), 'name': name})
    return jsonify(students_list)


@app.route('/get_absentees')
def get_absentees():
    """Return the list of students marked absent today (attendance == 1)."""
    df, err = read_and_index_df()
    if err:
        return jsonify({'error': err})

    date_col = datetime.now().strftime("%Y-%m-%d")
    if date_col not in df.columns:
        return jsonify([])

    absentees = df[df[date_col] == "1"]
    students_list = []
    for idx, row in absentees.iterrows():
        first = row.get('First Name') or row.get('First') or row.get('GivenName') or ''
        last = row.get('Last Name') or row.get('Last') or row.get('Surname') or ''
        name = f"{first} {last}".strip() or f"Student {idx}"
        students_list.append({'id': str(idx), 'name': name})
    return jsonify(students_list)


def update_excel_file_single(update):
    student_id = str(update.get('student_id', '')).strip()
    if not student_id:
        return {'success': False, 'message': 'Empty student_id provided.'}

    with save_lock:
        df, err = read_and_index_df()
        if err:
            return {'success': False, 'message': err}

        date_col = datetime.now().strftime("%Y-%m-%d")
        if date_col not in df.columns:
            df[date_col] = pd.NA

        if student_id not in df.index:
            return {'success': False, 'message': f"student_id {student_id} not found."}

        try:
            df.at[student_id, date_col] = update.get('attendance')
            df.to_excel(INPUT_EXCEL_FILE, index=False, engine='openpyxl')

            wb = load_workbook(INPUT_EXCEL_FILE)
            ws = wb.active
            col_idx = None
            for j, cell in enumerate(ws[1], start=1):
                if cell.value == date_col:
                    col_idx = j
                    break

            if col_idx:
                red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    val = cell.value
                    if val == 1 or str(val) == "1":
                        cell.fill = red_fill
                    elif val == 0.5 or str(val) == "0.5":
                        cell.fill = yellow_fill
                    else:
                        cell.fill = PatternFill(fill_type=None)

            wb.save(INPUT_EXCEL_FILE)
            return {'success': True, 'message': 'Attendance saved with formatting.'}
        except Exception as e:
            return {'success': False, 'message': f"Error saving Excel file: {e}"}


@app.route('/record_attendance', methods=['POST'])
def record_attendance():
    payload = request.get_json()
    if payload is None:
        return jsonify({'success': False, 'message': 'No JSON payload received.'})
    updates = payload if isinstance(payload, list) else [payload]
    results = [update_excel_file_single(upd) for upd in updates]
    all_ok = all(r.get('success') for r in results)
    messages = "; ".join(r.get('message', '') for r in results)
    return jsonify({'success': all_ok, 'message': messages})


def open_browser():
    time.sleep(1.5)  # Wait a bit for the server to start
    webbrowser.open('http://127.0.0.1:5000')


# --- Frontend ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>Attendance Taker</title>
<style>
  :root {
    --primary: #4361ee;
    --success: #4cc9f0;
    --danger: #f72585;
    --warning: #f9c74f;
    --light: #f8f9fa;
    --dark: #212529;
    --gray: #6c757d;
    --border-radius: 12px;
    --box-shadow: 0 10px 25px rgba(0,0,0,0.1);
    --transition: all 0.3s ease;
    --bg-present: linear-gradient(135deg, #d4edda 0%, #c3e6cb 100%);
    --bg-absent: linear-gradient(135deg, #f8d7da 0%, #f5c6cb 100%);
    --bg-excused: linear-gradient(135deg, #fff3cd 0%, #ffeaa7 100%);
    --bg-default: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
  }
  
  * {
    box-sizing: border-box;
    margin: 0;
    padding: 0;
  }
  
  html, body {
    height: 100%;
    transition: background 0.5s ease;
  }
  
  body {
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    display: flex;
    flex-direction: column;
    min-height: 100vh;
    margin: 0;
    background: var(--bg-default);
    padding: 20px;
    position: relative;
  }
  
  body.present-mode {
    background: var(--bg-present);
  }
  
  body.absent-mode {
    background: var(--bg-absent);
  }
  
  body.excused-mode {
    background: var(--bg-excused);
  }
  
  .container {
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    flex: 1;
    width: 100%;
  }
  
  .card {
    background: white;
    padding: 35px;
    border-radius: var(--border-radius);
    box-shadow: var(--box-shadow);
    width: 90%;
    max-width: 700px;
    text-align: center;
    transition: var(--transition);
  }
  
  .card:hover {
    box-shadow: 0 15px 35px rgba(0,0,0,0.15);
  }
  
  #name-display {
    font-size: 2.2rem;
    font-weight: 700;
    margin-bottom: 12px;
    color: var(--dark);
    min-height: 60px;
    display: flex;
    align-items: center;
    justify-content: center;
  }
  
  #status {
    margin-bottom: 25px;
    color: var(--gray);
    font-size: 1.1rem;
  }
  
  .btn-row {
    display: flex;
    gap: 15px;
    justify-content: center;
    flex-wrap: wrap;
    margin-bottom: 20px;
  }
  
  .btn {
    padding: 14px 24px;
    border-radius: var(--border-radius);
    border: none;
    font-weight: 600;
    cursor: pointer;
    color: white;
    min-width: 140px;
    transition: var(--transition);
    font-size: 1rem;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
  }
  
  .btn:hover {
    transform: translateY(-3px);
    box-shadow: 0 5px 15px rgba(0,0,0,0.2);
  }
  
  .btn:active {
    transform: translateY(0);
  }
  
  .present { background: var(--success); }
  .absent { background: var(--danger); }
  .excused { background: var(--warning); color: var(--dark); }
  
  .disabled {
    opacity: 0.6;
    pointer-events: none;
  }
  
  .progress-container {
    margin-top: 20px;
    margin-bottom: 10px;
    color: var(--gray);
    font-size: 1rem;
  }
  
  .page-footer {
    position: relative;
    width: 100%;
    margin-top: auto;
  }
  
  .recheck-btn {
    position: absolute;
    bottom: 0;
    right: 0;
    background: var(--primary);
    color: white;
    padding: 12px 20px;
    border: none;
    border-radius: var(--border-radius);
    cursor: pointer;
    font-weight: 600;
    transition: var(--transition);
    display: flex;
    align-items: center;
    gap: 8px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    margin: 20px;
  }
  
  .recheck-btn:hover {
    background: #3251d4;
    transform: translateY(-2px);
    box-shadow: 0 6px 16px rgba(0,0,0,0.2);
  }
  
  @media (max-width: 600px) {
    .card {
      padding: 25px 20px;
    }
    
    .btn-row {
      flex-direction: column;
      align-items: center;
    }
    
    .btn {
      width: 100%;
      max-width: 250px;
    }
    
    #name-display {
      font-size: 1.8rem;
    }
    
    .recheck-btn {
      position: relative;
      margin: 20px auto;
      right: auto;
      bottom: auto;
      width: calc(100% - 40px);
      max-width: 300px;
    }
  }
</style>
</head>
<body>
  <div class="container">
    <div class="card">
      <div id="name-display">Loading...</div>
      <div id="status">Press the button to record attendance</div>
      <div class="btn-row">
        <button id="present-btn" class="btn present">✅ Present</button>
        <button id="absent-btn"  class="btn absent">❌ Absent</button>
        <button id="excused-btn" class="btn excused">⏸️ Excused</button>
      </div>
      <div class="progress-container">Progress: <span id="progress">0 / 0</span></div>
    </div>
  </div>
  
  <div class="page-footer">
    <button id="recheck-btn" class="recheck-btn">🔄 Recheck Absentees</button>
  </div>

<script>
let students = [];
let idx = 0;
let recheckMode = false;

const nameEl = document.getElementById('name-display');
const statusEl = document.getElementById('status');
const progressEl = document.getElementById('progress');
const presentBtn = document.getElementById('present-btn');
const absentBtn = document.getElementById('absent-btn');
const excusedBtn = document.getElementById('excused-btn');
const recheckBtn = document.getElementById('recheck-btn');
const buttons = [presentBtn, absentBtn, excusedBtn];
const body = document.body;

function setButtonsDisabled(v){
  buttons.forEach(b => {
    if(v) b.classList.add('disabled'); else b.classList.remove('disabled');
    b.disabled = v;
  });
}

function setBackgroundColor(type) {
  // Remove any existing color classes
  body.classList.remove('present-mode', 'absent-mode', 'excused-mode');
  
  // Add the appropriate class based on the button pressed
  if (type === 'present') {
    body.classList.add('present-mode');
  } else if (type === 'absent') {
    body.classList.add('absent-mode');
  } else if (type === 'excused') {
    body.classList.add('excused-mode');
  }
  
  // Reset to default background after 1.5 seconds
  setTimeout(() => {
    body.classList.remove('present-mode', 'absent-mode', 'excused-mode');
  }, 1500);
}

async function fetchStudents(){
  try {
    const resp = await fetch('/get_students');
    const data = await resp.json();
    if(data.error){ nameEl.textContent = 'Error'; statusEl.textContent = data.error; return; }
    students = data; idx = 0; recheckMode = false;
    displayCurrent();
  } catch(e){ nameEl.textContent = 'Error loading'; statusEl.textContent = e.toString(); }
}

async function fetchAbsentees(){
  try {
    const resp = await fetch('/get_absentees');
    const data = await resp.json();
    if(data.error){ alert(data.error); return; }
    if(data.length === 0){ alert("No absentees found today ✅"); return; }
    students = data; idx = 0; recheckMode = true;
    displayCurrent();
  } catch(e){ alert("Error loading absentees: " + e.toString()); }
}

function displayCurrent(){
  if(idx >= students.length){
    if(recheckMode){ nameEl.textContent = 'Recheck complete ✅'; }
    else{ nameEl.textContent = 'Attendance complete ✅'; }
    statusEl.textContent = 'All done — Excel updated.';
    progressEl.textContent = `${students.length} / ${students.length}`;
    setButtonsDisabled(true);
    return;
  }
  const s = students[idx];
  nameEl.textContent = s.name;
  statusEl.textContent = 'Mark attendance';
  progressEl.textContent = `${idx + 1} / ${students.length}`;
  setButtonsDisabled(false);
}

async function mark(type){
  if(idx >= students.length) return;
  setButtonsDisabled(true);
  
  // Set background color based on button pressed
  setBackgroundColor(type);
  
  const s = students[idx];
  let value;
  if(type === 'present') value = 0;
  else if(type === 'absent') value = 1;
  else value = 0.5;

  statusEl.textContent = 'Saving...';
  const payload = { student_id: String(s.id), attendance: value };
  try {
    const resp = await fetch('/record_attendance', {
      method:'POST',
      headers:{ 'Content-Type':'application/json' },
      body: JSON.stringify(payload)
    });
    const res = await resp.json();
    if(!res.success){ statusEl.textContent = 'Save failed: ' + res.message; setButtonsDisabled(false); return; }
    idx++;
    setTimeout(displayCurrent, 300);
  } catch(err){
    statusEl.textContent = 'Network error'; setButtonsDisabled(false);
  }
}

presentBtn.addEventListener('click', () => mark('present'));
absentBtn.addEventListener('click', () => mark('absent'));
excusedBtn.addEventListener('click', () => mark('excused'));
recheckBtn.addEventListener('click', fetchAbsentees);

document.addEventListener('DOMContentLoaded', fetchStudents);
</script>
</body>
</html>
"""

if __name__ == '__main__':
    # Start the browser opening in a separate thread
    threading.Thread(target=open_browser).start()
    app.run(host='127.0.0.1', port=5000, debug=True)