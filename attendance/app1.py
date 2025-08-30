# app.py
# Attendance taker web app with recheck absentees feature.
# Requirements: flask, pandas, openpyxl
# pip install flask pandas openpyxl

import pandas as pd
from datetime import datetime
import threading
import os
import logging

from flask import Flask, render_template_string, request, jsonify
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')

app = Flask(__name__)

# --- Read Excel filename from text file ---
FILENAME_TXT = "input_filename.txt"
if os.path.exists(FILENAME_TXT):
    with open(FILENAME_TXT, "r", encoding="utf-8") as f:
        INPUT_EXCEL_FILE = f.read().strip()
else:
    INPUT_EXCEL_FILE = None

save_lock = threading.Lock()


def read_and_index_df():
    """Read the Excel file and return DataFrame indexed by OrgDefinedId."""
    if not INPUT_EXCEL_FILE or not os.path.exists(INPUT_EXCEL_FILE):
        return None, f"Input file '{INPUT_EXCEL_FILE}' not found."

    try:
        df = pd.read_excel(INPUT_EXCEL_FILE, engine='openpyxl', dtype=str)
    except Exception as e:
        return None, f"Error reading Excel file: {e}"

    if 'OrgDefinedId' not in df.columns:
        return None, "Required column 'OrgDefinedId' not found in the Excel file."

    df['OrgDefinedId'] = df['OrgDefinedId'].astype(str).str.strip()
    df.set_index('OrgDefinedId', inplace=True, drop=False)
    return df, None


@app.route('/')
def home():
    return HTML_TEMPLATE


@app.route('/get_students')
def get_students():
    df, err = read_and_index_df()
    if err:
        return jsonify({'error': err})

    students_list = []
    for idx, row in df.iterrows():
        first = row.get('First Name') or row.get('First') or row.get('GivenName') or ''
        last = row.get('Last Name') or row.get('Last') or row.get('Surname') or ''
        name = f"{first} {last}".strip() or f"Student {idx}"
        students_list.append({'id': str(idx), 'name': name})
    return jsonify(students_list)


@app.route('/get_absentees')
def get_absentees():
    """Return the list of students marked absent today (attendance == 1)."""
    df, err = read_and_index_df()
    if err:
        return jsonify({'error': err})

    date_col = datetime.now().strftime("%Y-%m-%d")
    if date_col not in df.columns:
        return jsonify([])

    absentees = df[df[date_col] == "1"]
    students_list = []
    for idx, row in absentees.iterrows():
        first = row.get('First Name') or row.get('First') or row.get('GivenName') or ''
        last = row.get('Last Name') or row.get('Last') or row.get('Surname') or ''
        name = f"{first} {last}".strip() or f"Student {idx}"
        students_list.append({'id': str(idx), 'name': name})
    return jsonify(students_list)


def update_excel_file_single(update):
    student_id = str(update.get('student_id', '')).strip()
    if not student_id:
        return {'success': False, 'message': 'Empty student_id provided.'}

    with save_lock:
        df, err = read_and_index_df()
        if err:
            return {'success': False, 'message': err}

        date_col = datetime.now().strftime("%Y-%m-%d")
        if date_col not in df.columns:
            df[date_col] = pd.NA

        if student_id not in df.index:
            return {'success': False, 'message': f"student_id {student_id} not found."}

        try:
            df.at[student_id, date_col] = update.get('attendance')
            df.to_excel(INPUT_EXCEL_FILE, index=False, engine='openpyxl')

            wb = load_workbook(INPUT_EXCEL_FILE)
            ws = wb.active
            col_idx = None
            for j, cell in enumerate(ws[1], start=1):
                if cell.value == date_col:
                    col_idx = j
                    break

            if col_idx:
                red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    val = cell.value
                    if val == 1 or str(val) == "1":
                        cell.fill = red_fill
                    elif val == 0.5 or str(val) == "0.5":
                        cell.fill = yellow_fill
                    else:
                        cell.fill = PatternFill(fill_type=None)

            wb.save(INPUT_EXCEL_FILE)
            return {'success': True, 'message': 'Attendance saved with formatting.'}
        except Exception as e:
            return {'success': False, 'message': f"Error saving Excel file: {e}"}


@app.route('/record_attendance', methods=['POST'])
def record_attendance():
    payload = request.get_json()
    if payload is None:
        return jsonify({'success': False, 'message': 'No JSON payload received.'})
    updates = payload if isinstance(payload, list) else [payload]
    results = [update_excel_file_single(upd) for upd in updates]
    all_ok = all(r.get('success') for r in results)
    messages = "; ".join(r.get('message', '') for r in results)
    return jsonify({'success': all_ok, 'message': messages})


# --- Frontend ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>Attendance Taker</title>
<style>
  body { font-family: Arial, Helvetica, sans-serif; display:flex; align-items:center; justify-content:center; height:100vh; margin:0; background:#f0f0f0; }
  .card { background:white; padding:28px; border-radius:12px; box-shadow:0 8px 30px rgba(0,0,0,0.08); width:90%; max-width:720px; text-align:center; position:relative; }
  #name-display { font-size:2rem; font-weight:700; margin-bottom:8px; }
  #status { margin-bottom:18px; color:#555; }
  .btn-row { display:flex; gap:12px; justify-content:center; }
  .btn { padding:12px 20px; border-radius:10px; border:none; font-weight:700; cursor:pointer; color:white; min-width:120px; }
  .present { background:#28a745; } .absent { background:#dc3545; } .excused { background:#f1c40f; color:#222; }
  .disabled { opacity:0.6; pointer-events:none; }
  .recheck-btn { position:absolute; bottom:15px; left:15px; background:#007bff; color:white; padding:10px 16px; border:none; border-radius:8px; cursor:pointer; font-weight:600; }
</style>
</head>
<body>
  <div class="card">
    <div id="name-display">Loading...</div>
    <div id="status">Press the button to record attendance</div>
    <div class="btn-row">
      <button id="present-btn" class="btn present">✅ Present</button>
      <button id="absent-btn"  class="btn absent">❌ Absent</button>
      <button id="excused-btn" class="btn excused">⏸️ Excused</button>
    </div>
    <div style="margin-top:14px; color:#888; font-size:0.9rem;">Progress: <span id="progress">0 / 0</span></div>
    <button id="recheck-btn" class="recheck-btn">🔄 Recheck Absentees</button>
  </div>

<script>
let students = [];
let idx = 0;
let recheckMode = false;

const nameEl = document.getElementById('name-display');
const statusEl = document.getElementById('status');
const progressEl = document.getElementById('progress');
const presentBtn = document.getElementById('present-btn');
const absentBtn = document.getElementById('absent-btn');
const excusedBtn = document.getElementById('excused-btn');
const recheckBtn = document.getElementById('recheck-btn');
const buttons = [presentBtn, absentBtn, excusedBtn];

function setButtonsDisabled(v){
  buttons.forEach(b => {
    if(v) b.classList.add('disabled'); else b.classList.remove('disabled');
    b.disabled = v;
  });
}

async function fetchStudents(){
  try {
    const resp = await fetch('/get_students');
    const data = await resp.json();
    if(data.error){ nameEl.textContent = 'Error'; statusEl.textContent = data.error; return; }
    students = data; idx = 0; recheckMode = false;
    displayCurrent();
  } catch(e){ nameEl.textContent = 'Error loading'; statusEl.textContent = e.toString(); }
}

async function fetchAbsentees(){
  try {
    const resp = await fetch('/get_absentees');
    const data = await resp.json();
    if(data.error){ alert(data.error); return; }
    if(data.length === 0){ alert("No absentees found today ✅"); return; }
    students = data; idx = 0; recheckMode = true;
    displayCurrent();
  } catch(e){ alert("Error loading absentees: " + e.toString()); }
}

function displayCurrent(){
  if(idx >= students.length){
    if(recheckMode){ nameEl.textContent = 'Recheck complete ✅'; }
    else{ nameEl.textContent = 'Attendance complete ✅'; }
    statusEl.textContent = 'All done — Excel updated.';
    progressEl.textContent = `${students.length} / ${students.length}`;
    setButtonsDisabled(true);
    return;
  }
  const s = students[idx];
  nameEl.textContent = s.name;
  statusEl.textContent = 'Mark attendance';
  progressEl.textContent = `${idx} / ${students.length}`;
  setButtonsDisabled(false);
}

async function mark(type){
  if(idx >= students.length) return;
  setButtonsDisabled(true);
  const s = students[idx];
  let value;
  if(type === 'present') value = 0;
  else if(type === 'absent') value = 1;
  else value = 0.5;

  statusEl.textContent = 'Saving...';
  const payload = { student_id: String(s.id), attendance: value };
  try {
    const resp = await fetch('/record_attendance', {
      method:'POST',
      headers:{ 'Content-Type':'application/json' },
      body: JSON.stringify(payload)
    });
    const res = await resp.json();
    if(!res.success){ statusEl.textContent = 'Save failed: ' + res.message; setButtonsDisabled(false); return; }
    idx++;
    setTimeout(displayCurrent, 300);
  } catch(err){
    statusEl.textContent = 'Network error'; setButtonsDisabled(false);
  }
}

presentBtn.addEventListener('click', () => mark('present'));
absentBtn.addEventListener('click', () => mark('absent'));
excusedBtn.addEventListener('click', () => mark('excused'));
recheckBtn.addEventListener('click', fetchAbsentees);

document.addEventListener('DOMContentLoaded', fetchStudents);
</script>
</body>
</html>
"""

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
